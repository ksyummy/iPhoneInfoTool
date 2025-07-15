import subprocess
import openpyxl
from datetime import datetime
import os

# 📁 Excel 파일명 (main.py와 같은 폴더에 위치)
FILEPATH = "iPhone_Info.xlsx"

# ✅ iPhone 정보 추출
def get_iphone_info():
    try:
        result = subprocess.check_output(['ideviceinfo'], encoding='utf-8')
    except Exception as e:
        return {"Error": f"iPhone 연결 또는 ideviceinfo 오류: {e}"}

    # 결과 파싱
    raw = {}
    for line in result.splitlines():
        if ":" in line:
            key, val = line.split(":", 1)
            raw[key.strip()] = val.strip()

    # 수동 입력 항목
    model = input("모델명을 입력하세요 (예: iPhone 15 Pro): ").strip()
    color = input("색상을 입력하세요 (예: 퍼플): ").strip()
    memory = input("메모리(용량)를 입력하세요 (예: 64G, 256G): ").strip()

    # 모델번호 + 지역코드 조합
    full_model_no = raw.get("ModelNumber", "") + raw.get("RegionInfo", "")

    # 정보 딕셔너리 반환
    info = {
        "모델": model,
        "색상": color,
        "메모리": memory,
        "모델번호": full_model_no,
        "IMEI1": raw.get("InternationalMobileEquipmentIdentity", ""),
        "IMEI2": raw.get("InternationalMobileEquipmentIdentity2", ""),
        "일련번호": raw.get("SerialNumber", ""),
        "EID": raw.get("EmbeddedIdentityDocument", "")
    }

    return info

# ✅ 실제 입력된 가장 가까운 빈 행 찾기
def find_next_empty_row(ws):
    row = 2  # 1행은 헤더
    while ws.cell(row=row, column=1).value not in (None, ""):
        row += 1
    return row

# ✅ Excel에 데이터 추가
def append_to_excel(filepath, data):
    if not os.path.exists(filepath):
        print(f"❌ 엑셀 파일을 찾을 수 없습니다: {filepath}")
        return

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    next_row = find_next_empty_row(ws)
    ws.cell(row=next_row, column=1).value = next_row - 1  # 번호 열 (A열)

    col_keys = ["모델", "색상", "메모리", "모델번호", "IMEI1", "IMEI2", "일련번호", "EID"]
    for i, key in enumerate(col_keys):
        ws.cell(row=next_row, column=i + 2).value = data.get(key, "")

    wb.save(filepath)
    print(f"✅ 정보가 엑셀에 저장되었습니다: {filepath}")

# ✅ 메인 실행 흐름
if __name__ == "__main__":
    info = get_iphone_info()
    if "Error" in info:
        print(info["Error"])
    else:
        append_to_excel(FILEPATH, info)

